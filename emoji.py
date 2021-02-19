# coding:utf-8
import requests
from bs4 import BeautifulSoup
import jieba
import os
import wordcloud
# from selenium import webdriver
import time
import random
import emoji
from openpyxl import Workbook
from openpyxl import load_workbook
proxies = {
    'http'  : 'socks5://127.0.0.1:10808',
    'https' : 'socks5://127.0.0.1:10808'
}
book = 'V8emojis.xlsx'
wb = load_workbook(book)
wb.save(book)
ws = wb.active
def gethtml(url):
    # options = webdriver.ChromeOptions()
    # options.add_argument('headless') #是否显示浏览器
    # driver = webdriver.Chrome(options=options)
    # driver.get(url)
    # time.sleep(1.0)
    # content = driver.page_source.encode('utf-8')
    content=requests.get(url)
    # print(content)
    soup=BeautifulSoup(content.text,"html.parser")
    # driver.quit()
    return soup
def findamount(url):
    page2=gethtml(url)
    tiezi=page2.find_all("span",attrs={"class":"red"})
    return int(tiezi[1].text)
def count(tie,counts):
    words = jieba.cut_for_search(tie)
    for word in words:
        if len(word) == 1:
            continue
        else:
            counts[word] = counts.get(word, 0) + 1
    return counts
def emoji(tie,emojis):
    for i in tie:
        # print(ord(i))
        if ord(i)<129686 and ord(i)>127743 :
            if i not in emojis:
                emojis[i]=1
            else:
                emojis[i]+=1
    return emojis
def tie(url):
    #含有访问
    link = "https://tieba.baidu.com/"+str(url)
    print(link)
    list=[]
    for i in range(1,findamount(link)):
        link2 =link+"?pn="+str(i)
        # print(link)
        p=gethtml(link2).find_all("div",attrs={"class":"d_post_content j_d_post_content"})
        for q in p:
            list.append(q.text.replace(" ",""))
    return list

def ties(url):
    #含有访问
    urllist=[]
    tiezis = gethtml(url).find_all('a', attrs={"class": "j_th_tit"})
    for tiezi in tiezis:
        # print(tiezi.get("title"))
        # print(tiezi.get("href"))
        urllist.append(tiezi.get("href"))
    return urllist
def clear(counts):
    for k in list(counts.keys()):
        val=counts[k]
        if val==1 or val==2 or val==3:
            counts.pop(k)
    return counts
def wordcloud(text):
    w = wordcloud.WordCloud(background_color='white')
    w.generate()
    w.to_file('pywcloud.png')
def pagelist(num,base_url):
    urllist = []
    for p in range(0, num):
        i = random.randint(0, 201)
        # i=4
        urllist.append(base_url + '?pn=' + str(50 * i))
        print('No' + str(i) + "正在筛选")
        print(urllist)
    return urllist
def load():
    print("正在读取旧数据")
    j = 2
    olddic={}
    ge1 = ws['A' + str(j)]
    ge2 = ws['B' + str(j)]
    mc = ge1.value
    olddic[mc] = ge2.value
    while mc != None:
        ge1 = ws['A' + str(j)]
        ge2 = ws['B' + str(j)]
        mc = ge1.value
        olddic[mc] = ge2.value
        j += 1
    print(olddic)
    print("读取旧数据完毕")
    return olddic
def sweep(newdic):
    # 去除字典空键
    newdic2={}
    for skey in newdic:
        if not newdic[skey] == None:
            newdic2[skey] = newdic[skey]
    return newdic2
def save(newdic2):
    o = 2
    for word in newdic2.keys():
        if ws["A" + str(o)] != None:
            ws["A" + str(o)] = word
            ws["B" + str(o)] = newdic2[word]
            o += 1
    wb.save(book)
if __name__ == '__main__':
    while 1:
        url = "https://tieba.baidu.com/f?kw=v&ie=utf-8"
        p=1
        # load()
        # emojis={'🐹': 3, '🔥': 136, '🌸': 12, '🤔': 30, '😭': 1771, '👊': 85, '😢': 28, '😋': 21, '🐴': 14, '🐶': 19, '😡': 42, '🧐': 4, '🤤': 37, '😅': 227, '🤓': 3, '😁': 24,
        #         '🤗': 80, '🤭': 26, '🍾': 246, '🙏': 16, '👋': 29, '🥱': 3, '😄': 20, '🤣': 96, '🙋': 5, '👀': 29, '🙇': 25, '😘': 5, '😃': 8, '🤐': 1, '😱': 57, '😓': 8, '😆': 23,
        #         '😨': 63, '🏳': 12, '😰': 15, '🐕': 1, '👆': 3, '🐒': 4, '😊': 42, '😈': 5, '🤮': 2, '🤩': 4, '💺': 1, '🍉': 1, '👴': 1, '🌚': 1, '💊': 11, '😵': 1, '🖕': 1, '🚬': 1,
        #         '👌': 4, '🏻': 10, '😜': 1, '💢': 1, '👹': 1, '🌛': 1, '🌀': 6, '👻': 3, '🍪': 1, '💖': 4, '😫': 1, '💕': 3, '💧': 7, '🌱': 1, '🌵': 1, '🎀': 5, '😍': 44, '😇': 21,
        #         '🥵': 5, '🐔': 2, '🙂': 2, '😑': 1, '😀': 1, '😬': 1, '😥': 3, '🥰': 26, '🐡': 1, '😂': 4, '👿': 4, '😙': 1, '🌽': 2, '🌿': 1, '🐵': 2, '💞': 3, '💋': 1, '🎃': 1,
        #         '🌈': 2, '😻': 1, '🎅': 1, '🍭': 1, '😚': 2, '💦': 1, '🏿': 57, '💃': 48, '👍': 2, '👩': 1, '🏃': 23, '💩': 1, '🥜': 1, '🏾': 6, '🌲': 2, '🙌': 1, '🤫': 7, '🤢': 3,
        #         '🐉': 1, '🐂': 1, '🥶': 1, '💥': 2, '🦈': 36, '🌶': 5, '😣': 1, '😉': 1, '😠': 1, '🐮': 2, '🍎': 1, '🤯': 1, '🎤': 2, '👽': 2, '🐰': 2, '🔪': 1, '😗': 3, '🐀': 1,
        #         '😸': 3, '🙅': 3}
        for link in pagelist(1,url):
            for url2 in ties(link):
                # time.sleep(5.0)
                try:
                    for i in tie(url2):
                        emojis=emoji(i,emojis)
                except:
                    save(emojis)
                    print("休息半个小时")
                    time.sleep(1800)
                    break
            print("ok" + str(p))
            print(emojis)
        save(emojis)
# print(sorted(emojis.items(), key=lambda d: d[1],reverse = True))





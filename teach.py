# coding:utf-8
import requests
from bs4 import BeautifulSoup
import jieba
import os
import wordcloud
# from selenium import webdriver
import time
import random
from openpyxl import Workbook
proxies = {
    'http'  : 'socks5://127.0.0.1:10808',
    'https' : 'socks5://127.0.0.1:10808'
}
def gethtml(url):
    # options = webdriver.ChromeOptions()
    # options.add_argument('headless') #是否显示浏览器
    # driver = webdriver.Chrome(options=options)
    # driver.get(url)
    # time.sleep(1.0)
    # content = driver.page_source.encode('utf-8')
    content=requests.get(url)
    # print(content)
    soup=BeautifulSoup(content.text,"html.parser")
    # driver.quit()
    return soup
def findamount(url):
    page2=gethtml(url)
    tiezi=page2.find_all("span",attrs={"class":"red"})
    return int(tiezi[1].text)
def count(tie,counts):
    words = jieba.cut_for_search(tie)
    for word in words:
        if len(word) == 1:
            continue
        else:
            counts[word] = counts.get(word, 0) + 1
    return counts
# https://tieba.baidu.com/p/7230184476?pn=3
def tie(url):
    link = "https://tieba.baidu.com/"+str(url)
    print(link)
    list=[]
    for i in range(1,findamount(link)):
        link2 =link+"?pn="+str(i)
        # print(link)
        p=gethtml(link2).find_all("div",attrs={"class":"d_post_content j_d_post_content"})
        for q in p:
            list.append(q.text.replace(" ",""))
    return list
def ties(url):
    urllist=[]
    tiezis = gethtml(url).find_all('a', attrs={"class": "j_th_tit"})
    for tiezi in tiezis:
        # print(tiezi.get("title"))
        # print(tiezi.get("href"))
        urllist.append(tiezi.get("href"))
    return urllist
def clear(counts):
    for k in list(counts.keys()):
        val=counts[k]
        if val==1 or val==2 or val==3:
            counts.pop(k)
    return counts
def wordcloud(text):
    w = wordcloud.WordCloud(background_color='white')
    w.generate()
    w.to_file('pywcloud.png')
def pagelist(num,base_url):
    urllist = []
    for p in range(0, num):
        i = random.randint(0, 201)
        # i=4
        urllist.append(base_url + '?pn=' + str(50 * i))
        print('No' + str(i) + "正在筛选")
        print(urllist)
    return urllist
counts=\
{'终于': 14, '回来': 59, '没人': 12, '回旋': 98, '木口': 16, '版本': 37, '不能': 102, '接受': 13, '几个': 32, '一样': 63, 'v8': 58, '哪里': 13, '杏奴': 473, '不到': 19, '什么': 272, '应该': 30, '只要': 18, '感情': 11, '觉得': 40, 'holo': 63, '真的': 89, '这样': 51, '自己': 97, '炒作': 63, '没有': 110, '感觉': 24, '他们': 60, '原谅': 10, '可以': 112, '问题': 25, '不会': 25, '现在': 83, '不是': 180, '而且': 11, '一眼': 24, '只有': 23, '努力': 16, '摆烂': 14, '就是': 131, '...': 26, '麻麻': 19, '以后': 15, '资源': 9, '可惜': 10, '确实': 42, '能力': 12, '一直': 56, '之前': 17, 'apex': 8, '喜欢': 137, '游戏': 18, '这个': 73, '不看': 51, '直播': 99, '时间': 27, '有点': 33, '不住': 42, '不了': 21, '以前': 24, '这么': 48, '意思': 14, '干嘛': 5, '真是': 18, '营业': 12, '段时间': 7, '希望': 31, '地方': 16, '如果': 58, '尽量': 12, '赶紧': 10, '大家': 53, '看到': 30, '一下': 76, '出来': 22, '发情': 7, '为什么': 51, '还要': 15, '罢了': 21, '的话': 29, '世界': 25, '其实': 26, '字母': 22, '连体': 6, '今日': 11, '入土': 23, '私信': 18, '想想': 17, '小林': 28, '马上': 12, '以为': 23, '嘉然': 28, '我们': 100, '怎么': 87, '天狗': 24, '胜利': 10, '日本': 20, '支持': 305, '很难': 27, '抗抗': 7, '时候': 85, 'LOL': 6, '原来': 8, '还是': 69, '粉丝': 27, '你们': 83, '金盾': 7, '因为': 49, '可能': 62, '小姐': 26, '第一': 33, '一个': 115, '没想': 18, '想到': 23, '没想到': 18, '起来': 19, 'hololive': 35, '暴力': 10, '继续': 17, '一次': 30, '回复': 22, '离开': 17, '四个': 15, '国内': 15, '已经': 49, '猪圈': 13, '芜狐': 29, 'washoi': 6, 'nanodesu': 6, '迷迭': 6, '可愛': 6, 'neeee': 6, 'moooo': 6, 'muA': 6, 'dd': 6, '斩首': 6, '敲理': 6, 'moneymoneymoney': 6, 'Pekopeko': 6, '牛顿': 20, '一边': 11, '结果': 16, '哈哈': 39, '哈哈哈': 18, '慢慢': 11, '泥哥': 20, '最好': 15, '飞龙': 4, '开心': 19, '这种': 21, '心情': 5, '复杂': 8, '意识': 11, 'cn': 7, '腾达': 4, '飞黄腾达': 4, '但是': 42, '当初': 4, '大佐': 8, '白上': 28, '吹雪': 28, '可爱': 56, '打卡': 130, '九节': 7, '九节鞭': 7, '今天': 80, '睡觉': 10, '帖子': 11, '每天': 21, '嘎嘎': 94, '不错': 9, '早上': 14, '知道': 67, '拉夫': 23, 'buff': 29, 'lulu': 41, '百合': 20, '声音': 8, '还有': 48, '迟早': 11, '罕见': 9, 'yhm': 7, '不过': 35, '下去': 16, '起床': 9, '不然': 12, '根本': 9, '晚上': 14, '没法': 5, '安心': 6, '阴阳': 16, '怪气': 16, '阴阳怪气': 16, '怎么样': 16, '一切': 54, '灵魂': 5, '不管': 11, '图片': 46, '看看': 20, '评论': 19, '严重': 8, '今晚': 9, '进去': 8, 'mea': 80, '麻将': 5, '看不懂': 23, '反转': 11, 'run': 5, '明天': 17, 'sc': 6, '后台': 8, '杂谈': 6, '打游戏': 4, '女人': 15, 'Mea': 7, '彗酱': 8, 'vtuber': 38, '自由': 30, '讨论': 15, '或者': 22, '其他': 24, '本来': 10, '然而': 19, '也好': 10, '斧子': 133, '之类': 26, '甚至': 16, '使用': 36, '处理': 33, '了解': 10, '这些': 24, '帽子': 9, '细则': 9, '情况': 23, '不要': 37, '完全': 13, '气氛': 14, '任何': 25, '行为': 11, '艾特': 15, '说明': 12, '谢谢': 15, '理解': 28, '荷花': 85, '孝子': 52, '4000': 356, '早该': 39, '管管': 28, '指导': 31, '坚持': 29, '凡是': 23, '永远': 42, '正义': 45, '越来': 20, '越来越': 20, '字词': 46, '过来': 12, '为了': 18, '结束': 8, '温柔': 32, '妈妈': 33, '决定': 9, '遛狗': 19, '十月': 28, '妄想': 25, '有人': 31, '那个': 18, '经典': 11, '单纯': 10, '只能': 21, '这下': 27, '有些': 16, '变成': 22, '乐子': 12, '直接': 17, '不得': 18, '记忆': 9, '不多': 33, '差不多': 33, '家伙': 10, '好家伙': 8, '各位': 7, '一定': 13, '老婆': 158, '下来': 14, '经常': 17, '天天': 16, '同学': 6, '我俩': 24, '一起': 39, '吃饭': 6, '班上': 15, '特别': 9, '朋友': 24, '很少': 9, '面前': 7, '在意': 7, '满足': 13, '一点': 21, '然后': 41, '最后': 29, '回家': 11, '学校': 19, '每个': 20, '月底': 10, '事情': 17, '发现': 14, '他妹': 18, '告诉': 15, '月初': 7, '肯定': 40, '于是': 7, '当时': 16, '样子': 14, '很多': 20, '所有': 16, '这件': 9, '虽然': 24, '半天': 7, '反正': 10, '后面': 9, '发生': 10, '就算': 14, '那样': 9, '多少': 9, '恶心': 42, '点击': 26, '展开': 28, '查看': 26, '完整': 30, '是不是': 19, '引流': 9, '除了': 13, '十全': 5, '十全十美': 5, '看个': 5, '气要': 5, '那么': 20, '航天': 11, '实验': 6, '研究': 11, '循环': 9, '肯尼': 6, '中心': 7, '肯尼迪': 6, '肯尼迪航天中心': 6, '美元': 4, '如此': 12, '进来': 7, '居然': 5, '地球': 7, '自转': 4, '探测': 8, '土星': 6, '系统': 7, '测器': 6, '探测器': 6, '12': 7, '关系': 13, '物质': 5, '好好': 17, '无法': 18, 'meaqua': 9, '物理': 17, '三大': 17, '定律': 17, '发声': 4, '最终': 6, '异色': 4, '脸庞': 5, '想要': 23, '听到': 6, '歌声': 4, '喜悦': 4, '突然': 15, '看着': 19, '泛起': 4, '里面': 5, '不敢': 11, '一声': 7, '惊喜': 5, '扑闪': 4, '太多': 8, '身边': 8, '此生': 4, '生命': 9, '女生': 7, 'asoul': 12, '所谓': 14, '所以': 16, '温暖': 7, '家庭': 8, '大家庭': 7, '时代': 12, '一种': 696, '偶像': 25, '叫做': 691, '时乃空': 24, '软萌': 20, '萝卜': 24, '完美': 31, '星街': 24, '彗星': 30, '冒失': 20, '赤井心': 24, '活泼': 20, '巫女': 24, '甜蜜': 20, '夜空': 24, '梅露': 24, '仁爱': 20, '大神': 24, '混沌': 20, '夏色': 24, '阿库娅': 25, '自慢': 20, '蠢萌': 20, '百鬼绫目': 24, '愈月巧': 23, '元气': 20, '大空': 24, '池面': 20, '神沁音': 24, '傲娇': 20, '兔田佩': 24, '克拉': 24, '奔放': 20, '润羽': 24, '露西': 24, '西亚': 24, '露西亚': 24, '宁静': 20, '不知': 27, '火芙蕾雅': 24, '广阔': 20, '白银': 25, '诺艾尔': 24, '调皮': 20, '宝钟': 24, '玛琳': 24, '天音': 24, '彼方': 24, '桐生': 25, '可可': 25, '善良': 27, '角卷': 24, '绵芽': 24, '反差': 20, '常暗': 24, '腹黑': 20, '姬森璐娜': 24, '喜爱': 20, '美好': 25, 'Hololive': 20, 'crew': 62, '存在': 10, '开团': 10, 'V8': 17, 'kiss': 13, '认识': 24, '账号': 5, '那天': 7, '团长': 10, '需要': 20, '开始': 20, '半夜': 5, '几点': 5, '8426': 5, '造谣': 6, '疯狂': 12, '泥潭': 7, '猫雷': 333, '后来': 8, '一级': 4, '对线': 5, '之后': 13, '关注': 26, '到底': 12, '作为': 11, '梁木': 9, '我点': 4, 'NGA': 6, '别人': 15, '我爱你': 335, '记得': 8, '清楚': 6, '8u': 30, '看杏': 20, '想起': 4, '一天': 9, '互相': 5, '阿梓': 8, '破防': 11, '好事': 12, '绝对': 5, '会长': 4, '有时': 10, '有时候': 9, '只是': 30, '关心': 6, '滚出': 7, '这里': 9, '阿夸': 69, '匆匆': 4, '默默': 4, '守护': 4, '害怕': 8, '非常': 14, '名气': 5, '心思': 4, '唱歌': 12, '伤害': 6, 'as': 27, 'rx': 19, '取关': 4, '苛责': 6, '女孩': 21, '直播间': 33, '欢快': 6, '弹幕': 23, '容易': 11, '可是': 8, '互动': 12, '那些': 6, '快乐': 8, '发病': 16, '她们': 8, '犯错': 6, '普通': 5, '东西': 19, '明明': 7, '学着': 4, '男人': 17, '控制': 5, '好像': 6, '小红': 12, '正常': 23, '抱歉': 5, '深渊': 4, '收租': 7, '小时': 11, '哔哩': 32, '播到': 18, '四点': 8, 'Vtuber': 4, '老师': 5, '新月': 26, 'mmr': 18, '司马': 4, '大司马': 4, '忏悔': 5, '似拟': 6, '六边': 4, '六边形': 4, '死角': 4, '战士': 5, '行星': 4, '而已': 11, '黑暗': 10, '宇宙': 5, '闪耀': 4, '即便': 8, '即便如此': 4, '直到': 5, '紫联': 14, '诗音': 14, '矛盾': 29, '转冲': 24, '杏仁': 12, '好话': 10, '全给': 10, '人生': 10, '大意': 9, 'lolo': 6, '丝粉': 6, '乖乖': 5, '是己': 6, '装假': 6, '不惯': 12, '看不惯': 12, '别看': 12, '杏瘾': 7, '啊啊啊': 12, '会战': 7, 'r9': 5, '冲刺': 9, '安慰': 5, '贝拉': 11, '观众': 12, '主播': 10, '效果': 9, 'AS': 9, '切割': 19, '早点': 14, '彼此': 15, '真实': 5, '虚假': 6, '收到': 4, '一帮': 5, '老鼠': 8, '鄙视': 5, '说到底': 5, '岚宝': 10, '运气': 5, '岚宝涨': 5, '粉靠': 5, '整活': 6, '是因为': 7, '好听': 6, '披露': 8, '贵物': 9, '缝合': 10, '抗压': 7, '裤裆': 12, 'ybb': 5, '脆脆': 4, '海子': 5, '拿下': 15, '执行': 5, '标记': 7, '中国': 10, '长大': 4, '快快': 28, '大脑': 5, '牵引': 5, '歪歪': 12, '着迷': 4, '荷尔': 4, '荷尔蒙': 4, '陷阱': 4, '逃离': 4, '歌姬': 7, '橄榄': 7, '毕业': 5, '自卑': 5, '抑郁': 6, '一些': 4, '难受': 4, '郭楠': 4, '女性': 6, '常人': 4, '正常人': 4, '母人': 5, '男女': 4, '总是': 4, '意志': 4, '操纵': 4, '操纵者': 4, '前路': 4, '一丝': 4, '光明': 4, '来来': 6, '不说': 32, '晚安': 32, '蓝牌': 6, '尸体': 9, '绅宝': 12, '绅绅': 7, '好玩': 4, '卧槽': 4, '哈人': 4, '你好': 6, '对不起': 8, '求求': 14, '列表': 4, '一条': 4, 'vx': 4, '加个': 4, '生活': 4, '内战': 4}
url = "https://tieba.baidu.com/f?kw=v&ie=utf-8"
p=1
for link in pagelist(1,url):
    for url2 in ties(link):
        time.sleep(5.0)
        for i in tie(url2):
            counts=count(i,counts)
            # print(counts)
        print("ok" + str(p))
        p+=1
        counts = clear(counts)
        print(counts)
print(sorted(counts.items(), key=lambda d: d[1],reverse = True))
wb=Workbook()
ws=wb.active
i=1
for word in counts.keys():
    ws["A"+str(i)]=word
    ws["B"+str(i)]=counts[word]
    i += 1
wb.save(r"words.xlsx")







